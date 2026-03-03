"""
================================================================================
Spatial Block Cross-Validation + Random Forest Counterfactual
Paddy vs Upland: SOC and SIC comparison at two depth intervals
================================================================================
Design principles:
 - Spatial blocks (100 km grid) to prevent spatial autocorrelation leakage
 - Out-of-fold (OOF) prediction for unbiased CV metrics
 - Counterfactual Δ = Pred(Upland) − Pred(Paddy) on paddy sites (OOF only)
 - Block bootstrap for 95% CI (robust to spatial clustering)
 - log1p transform for right-skewed carbon data
 - mtry ≈ p/3 (recommended for regression RF)
 - Reports CV R² and RMSE in BOTH log and original scale
 - Nature double-column publication figure (183 mm wide)
================================================================================
"""

import warnings
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
from matplotlib import rcParams
from matplotlib.ticker import AutoMinorLocator
from scipy.stats import mannwhitneyu
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import r2_score, mean_squared_error
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Typography (Nature style) ────────────────────────────────────────────────
rcParams.update({
    'font.family'      : 'Arial',
    'font.size'        : 8,
    'axes.labelsize'   : 9,
    'axes.titlesize'   : 9.5,
    'xtick.labelsize'  : 8,
    'ytick.labelsize'  : 8,
    'legend.fontsize'  : 8,
    'axes.linewidth'   : 0.8,
    'xtick.major.width': 0.6,
    'ytick.major.width': 0.6,
    'xtick.major.size' : 3.0,
    'ytick.major.size' : 3.0,
    'xtick.direction'  : 'in',
    'ytick.direction'  : 'in',
    'pdf.fonttype'     : 42,
    'ps.fonttype'      : 42,
})

# ════════════════════════════════════════════════════════════════════════════════
# 0. PARAMETERS
# ════════════════════════════════════════════════════════════════════════════════
DATA_PATH  = r'C:\Users\Admin\Desktop\MS\原始数据\China_profile_dataset_TC_added.xlsx'
OUT_DIR    = r'C:\Users\Admin\Desktop\MS\修改后代码与图片'
GRID_KM    = 100       # spatial block size (km)
K_FOLDS    = 5         # spatial CV folds
N_REPS     = 5         # repetitions (different fold assignments)
N_TREES    = 500       # RF trees
MIN_LEAF   = 5         # minimum samples per leaf
BOOT_B     = 800       # block bootstrap resamples
SEED       = 2026

# Feature columns (environmental predictors)
ENV_COLS = ['BD', 'DEM', 'NDVI', 'MAT', 'MAP', 'PET', 'AI', 'pH', 'Sand', 'Silt', 'Clay']

print("=" * 70)
print("Spatial Block CV + RF Counterfactual Model")
print("=" * 70)

# ════════════════════════════════════════════════════════════════════════════════
# 1. LOAD & PREPROCESS DATA
# ════════════════════════════════════════════════════════════════════════════════
print("\n[1] Loading data...")
df_raw = pd.read_excel(DATA_PATH)
print(f"    Loaded: {len(df_raw):,} rows × {df_raw.shape[1]} cols")

# Robust column matching (case-insensitive, strip whitespace)
def find_col(df, pattern):
    """Find column by partial case-insensitive match."""
    pattern_l = pattern.lower().replace(' ', '_')
    for c in df.columns:
        cn = c.strip().lower().replace(' ', '_').replace('-', '_')
        if pattern_l in cn or cn in pattern_l:
            return c
    # fallback: substring
    for c in df.columns:
        cn = c.strip().lower()
        if pattern_l[:6] in cn:
            return c
    raise KeyError(f"Cannot find column matching '{pattern}'. Available: {list(df.columns)[:20]}")

lu_col  = find_col(df_raw, 'landuse_tol1km_2019')
lon_col = find_col(df_raw, 'longitude')
lat_col = find_col(df_raw, 'latitude')
up_col  = find_col(df_raw, 'upper_dept')
lo_col  = find_col(df_raw, 'lower_dept')

# Resolve env columns
env_found = []
env_labels = []
for ec in ENV_COLS:
    try:
        c = find_col(df_raw, ec)
        env_found.append(c)
        env_labels.append(ec)
    except KeyError:
        print(f"    Warning: predictor '{ec}' not found, skipping")

print(f"    Predictors found: {env_labels}")

# Land use classification
def classify_lu(x):
    x = str(x).lower()
    if 'paddy' in x: return 'Paddy'
    if 'upland' in x: return 'Upland'
    return np.nan

# Depth grouping (mid-point of layer)
def depth_group(up, lo):
    try:
        up, lo = float(up), float(lo)
        mid = (up + lo) / 2
        if mid <= 30:   return '0–30 cm'
        if mid <= 100:  return '30–100 cm'
    except:
        pass
    return np.nan

df = df_raw.copy()
df['lu']   = df[lu_col].apply(classify_lu)
df['upper'] = pd.to_numeric(df[up_col], errors='coerce')
df['lower'] = pd.to_numeric(df[lo_col], errors='coerce')
df['depth'] = df.apply(lambda r: depth_group(r['upper'], r['lower']), axis=1)
df['lon']   = pd.to_numeric(df[lon_col], errors='coerce')
df['lat']   = pd.to_numeric(df[lat_col], errors='coerce')

# Numeric env features
for ec, el in zip(env_found, env_labels):
    df[f'feat_{el}'] = pd.to_numeric(df[ec], errors='coerce')
feat_cols = [f'feat_{el}' for el in env_labels]

# Spatial block ID (100-km grid in degrees ≈ 0.9°)
# Convert to Web Mercator equivalents (simple approximation)
deg_per_km = 1.0 / 111.0
BLOCK_DEG = GRID_KM * deg_per_km
df['block_id'] = (
    (df['lon'] / BLOCK_DEG).apply(np.floor).astype(str) + '_' +
    (df['lat'] / BLOCK_DEG).apply(np.floor).astype(str)
)

# Filter valid rows
df_base = df.dropna(subset=['lu', 'lon', 'lat', 'depth', 'block_id']).copy()
df_base = df_base[df_base['lu'].isin(['Paddy', 'Upland'])].copy()

print(f"    Valid rows after filtering: {len(df_base):,}")
print(f"    Land use counts:\n{df_base['lu'].value_counts().to_string()}")
print(f"    Depth groups:\n{df_base['depth'].value_counts().to_string()}")
print(f"    Unique spatial blocks: {df_base['block_id'].nunique()}")

# ════════════════════════════════════════════════════════════════════════════════
# 2. HELPER FUNCTIONS
# ════════════════════════════════════════════════════════════════════════════════

def make_block_folds(block_ids, k=5, seed=0):
    """Assign each block to a fold, return per-sample fold labels."""
    rng = np.random.RandomState(seed)
    unique_blocks = list(set(block_ids.dropna()))
    rng.shuffle(unique_blocks)
    fold_map = {b: (i % k) + 1 for i, b in enumerate(unique_blocks)}
    return block_ids.map(fold_map)


def block_bootstrap_ci(delta, block_ids, B=800, seed=2026):
    """
    Block bootstrap CI for mean(delta).
    Takes mean per block first to avoid large-block dominance.
    """
    mask = ~(np.isnan(delta) | block_ids.isna())
    delta_ok = delta[mask]
    blk_ok   = block_ids[mask]
    blk_means = delta_ok.groupby(blk_ok).mean().values
    m = len(blk_means)
    if m < 5:
        return dict(mean=np.nanmean(delta_ok), lo=np.nan, hi=np.nan, n_blocks=m)
    rng = np.random.RandomState(seed)
    boot_means = [rng.choice(blk_means, size=m, replace=True).mean() for _ in range(B)]
    return dict(
        mean     = float(blk_means.mean()),
        lo       = float(np.percentile(boot_means, 2.5)),
        hi       = float(np.percentile(boot_means, 97.5)),
        n_blocks = m
    )


def compute_r2_rmse(obs, pred):
    mask = ~(np.isnan(obs) | np.isnan(pred))
    if mask.sum() < 5:
        return np.nan, np.nan
    o, p = obs[mask], pred[mask]
    r2   = 1 - np.sum((o - p)**2) / np.sum((o - np.mean(o))**2)
    rmse = np.sqrt(np.mean((o - p)**2))
    return float(r2), float(rmse)


# ════════════════════════════════════════════════════════════════════════════════
# 3. CORE: SPATIAL BLOCK CV COUNTERFACTUAL
# ════════════════════════════════════════════════════════════════════════════════

def run_spatialcv_rf_counterfactual(df_all, y_col, response_label, depth_label):
    """
    Spatial block CV RF counterfactual:
      Δ = Pred(Upland) − Pred(Paddy) on paddy sites (out-of-fold only)

    Returns performance metrics and bootstrap CI for Δ.
    """
    # ── Subset
    sub = df_all[df_all['depth'] == depth_label].copy()
    sub = sub.dropna(subset=[y_col] + feat_cols + ['lu', 'block_id', 'lon', 'lat'])
    sub = sub[sub[y_col] >= 0].copy()
    sub = sub[sub['lu'].isin(['Paddy', 'Upland'])].copy()

    n_total  = len(sub)
    n_paddy  = (sub['lu'] == 'Paddy').sum()
    n_upland = (sub['lu'] == 'Upland').sum()
    n_blocks = sub['block_id'].nunique()

    label_str = f"{response_label} | {depth_label}"
    print(f"\n  [{label_str}]  n={n_total}  paddy={n_paddy}  upland={n_upland}  blocks={n_blocks}")

    if n_total < 200 or min(n_paddy, n_upland) < 20:
        print(f"    ✗ Insufficient data, skipping.")
        return None

    sub = sub.reset_index(drop=True)
    sub['y_log']  = np.log1p(sub[y_col].values)
    sub['lu_bin'] = (sub['lu'] == 'Paddy').astype(int)   # 1=paddy, 0=upland

    predictors = feat_cols + ['lu_bin']
    n_pred = len(predictors)
    mtry   = max(3, n_pred // 3)   # p/3 for regression RF

    # Storage across reps
    idx_paddy = sub.index[sub['lu'] == 'Paddy'].tolist()
    oof_log   = np.full((n_total, N_REPS), np.nan)
    delta_mat = np.full((len(idx_paddy), N_REPS), np.nan)
    paddy_pos = {idx: i for i, idx in enumerate(idx_paddy)}

    for rep in range(N_REPS):
        fold_ids = make_block_folds(sub['block_id'], k=K_FOLDS, seed=SEED + rep)
        sub[f'fold_{rep}'] = fold_ids

        for fold in range(1, K_FOLDS + 1):
            tr_mask = fold_ids != fold
            te_mask = fold_ids == fold

            tr_data = sub[tr_mask]
            te_data = sub[te_mask]

            # Guard: training must have both land uses
            if tr_data['lu'].nunique() < 2:
                continue

            X_tr = tr_data[predictors].values
            y_tr = tr_data['y_log'].values

            rf = RandomForestRegressor(
                n_estimators=N_TREES,
                max_features=mtry,
                min_samples_leaf=MIN_LEAF,
                n_jobs=-1,
                random_state=SEED
            )
            rf.fit(X_tr, y_tr)

            # OOF predictions (log scale)
            X_te = te_data[predictors].values
            oof_log[te_mask, rep] = rf.predict(X_te)

            # Counterfactual Δ: paddy sites in this test fold
            te_paddy_idx = te_data.index[te_data['lu'] == 'Paddy'].tolist()
            if not te_paddy_idx:
                continue

            te_p = te_data.loc[te_paddy_idx, predictors].copy()
            te_u = te_data.loc[te_paddy_idx, predictors].copy()
            te_p['lu_bin'] = 1   # keep as paddy
            te_u['lu_bin'] = 0   # counterfactual: switch to upland

            pred_p = np.expm1(rf.predict(te_p.values))
            pred_u = np.expm1(rf.predict(te_u.values))
            delta  = pred_u - pred_p   # Δ = Upland − Paddy

            for i, idx in enumerate(te_paddy_idx):
                pos = paddy_pos[idx]
                delta_mat[pos, rep] = delta[i]

        print(f"    Rep {rep+1}/{N_REPS} done", end='\r')
    print()

    # ── CV performance (log scale) ──────────────────────────────────────────
    obs_log  = sub['y_log'].values
    pred_log = np.nanmean(oof_log, axis=1)
    r2_log, rmse_log = compute_r2_rmse(obs_log, pred_log)

    # CV performance (original scale)
    obs_orig  = sub[y_col].values
    pred_orig = np.expm1(pred_log)
    r2_orig, rmse_orig = compute_r2_rmse(obs_orig, pred_orig)

    # ── Block bootstrap CI for Δ ─────────────────────────────────────────────
    delta_pt = pd.Series(np.nanmean(delta_mat, axis=1),
                         index=idx_paddy)
    paddy_sub = sub.loc[idx_paddy]
    ci = block_bootstrap_ci(delta_pt, paddy_sub['block_id'],
                            B=BOOT_B, seed=SEED)

    sig = ('Increase' if ci['lo'] > 0 else
           'Decrease' if ci['hi'] < 0 else 'ns')

    print(f"    CV R2(log)={r2_log:.3f}  RMSE(log)={rmse_log:.3f}")
    print(f"    CV R2(orig)={r2_orig:.3f}  RMSE(orig)={rmse_orig:.3f} g/kg")
    print(f"    Delta = {ci['mean']:+.3f} g/kg  95%CI[{ci['lo']:+.3f}, {ci['hi']:+.3f}]  {sig}")

    return dict(
        Response      = response_label,
        Depth         = depth_label,
        n_total       = n_total,
        n_paddy       = n_paddy,
        n_upland      = n_upland,
        n_blocks      = n_blocks,
        n_predictors  = n_pred,
        mtry          = mtry,
        n_trees       = N_TREES,
        k_folds       = K_FOLDS,
        n_reps        = N_REPS,
        CV_R2_log     = round(r2_log,  3),
        CV_RMSE_log   = round(rmse_log, 3),
        CV_R2_orig    = round(r2_orig,  3),
        CV_RMSE_orig  = round(rmse_orig, 3),
        Mean_Delta    = round(ci['mean'], 4),
        CI_lo95       = round(ci['lo'],  4),
        CI_hi95       = round(ci['hi'],  4),
        Significance  = sig,
        n_boot_blocks = ci['n_blocks'],
    )


# ════════════════════════════════════════════════════════════════════════════════
# 4. RUN MODEL FOR SOC AND SIC × TWO DEPTHS
# ════════════════════════════════════════════════════════════════════════════════
print("\n[2] Running models...")

targets = [
    ('SOC', 'SOC'),
    ('sic_g_kg_ovl', 'SIC'),
    ('TC', 'TC'),
]
depths = ['0–30 cm', '30–100 cm']

# locate SOC / SIC / TC columns
soc_col = find_col(df_raw, 'SOC')
sic_col = find_col(df_raw, 'sic_g_kg_ovl')
tc_col  = find_col(df_raw, 'TC')

# Add y columns to working df
df_base['SOC'] = pd.to_numeric(df_raw.loc[df_base.index, soc_col] if soc_col in df_raw.columns
                                else df_raw[soc_col], errors='coerce')
df_base['SIC'] = pd.to_numeric(df_raw.loc[df_base.index, sic_col] if sic_col in df_raw.columns
                                else df_raw[sic_col], errors='coerce')
df_base['TC']  = pd.to_numeric(df_raw.loc[df_base.index, tc_col] if tc_col in df_raw.columns
                                else df_raw[tc_col], errors='coerce')

results = []
for y_key, label in targets:
    for dep in depths:
        res = run_spatialcv_rf_counterfactual(df_base, y_key, label, dep)
        if res is not None:
            results.append(res)

results_df = pd.DataFrame(results)
print("\n" + "="*70)
print(results_df[['Response','Depth','n_total','n_paddy','CV_R2_orig','CV_RMSE_orig',
                   'Mean_Delta','CI_lo95','CI_hi95','Significance']].to_string(index=False))

# ════════════════════════════════════════════════════════════════════════════════
# 5. EXPORT RESULTS TABLE (EXCEL, FORMATTED)
# ════════════════════════════════════════════════════════════════════════════════
print("\n[3] Saving results table...")

# --- Pretty column names for table ---
col_rename = {
    'Response'     : 'Variable',
    'Depth'        : 'Depth interval',
    'n_total'      : 'n (total)',
    'n_paddy'      : 'n (Paddy)',
    'n_upland'     : 'n (Upland)',
    'n_blocks'     : 'Spatial blocks',
    'n_predictors' : 'No. predictors',
    'mtry'         : 'mtry',
    'n_trees'      : 'No. trees',
    'k_folds'      : 'CV folds',
    'n_reps'       : 'CV repetitions',
    'CV_R2_log'    : 'CV R² (log scale)',
    'CV_RMSE_log'  : 'CV RMSE (log scale)',
    'CV_R2_orig'   : 'CV R² (original)',
    'CV_RMSE_orig' : 'CV RMSE (g kg⁻¹)',
    'Mean_Delta'   : 'Δ (g kg⁻¹)',
    'CI_lo95'      : '95% CI lower',
    'CI_hi95'      : '95% CI upper',
    'Significance' : 'Significance',
    'n_boot_blocks': 'Bootstrap blocks',
}
table_df = results_df.rename(columns=col_rename)

# Write styled Excel
xlsx_path = f'{OUT_DIR}\\SpatialCV_RF_Results.xlsx'
wb = openpyxl.Workbook()

# ── Sheet 1: Summary table ──────────────────────────────────────────────────
ws = wb.active
ws.title = 'Results'

header_fill = PatternFill('solid', fgColor='1F3864')
alt_fill    = PatternFill('solid', fgColor='DCE6F1')
hdr_font    = Font(name='Arial', size=9, bold=True, color='FFFFFF')
cell_font   = Font(name='Arial', size=9)
ctr_align   = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)

cols = list(table_df.columns)
for ci_col, col_name in enumerate(cols, 1):
    cell = ws.cell(row=1, column=ci_col, value=col_name)
    cell.font      = hdr_font
    cell.fill      = header_fill
    cell.alignment = ctr_align
    cell.border    = thin_border

for ri, row in enumerate(table_df.itertuples(index=False), 2):
    fill = alt_fill if ri % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
    for ci_col, val in enumerate(row, 1):
        val_out = val
        # Round floats
        if isinstance(val, float):
            val_out = round(val, 3)
        cell = ws.cell(row=ri, column=ci_col, value=val_out)
        cell.font      = cell_font
        cell.alignment = ctr_align
        cell.fill      = fill
        cell.border    = thin_border
        # Colour significance cell
        if cols[ci_col-1] == 'Significance':
            if val == 'Increase':
                cell.font = Font(name='Arial', size=9, bold=True, color='C00000')
            elif val == 'Decrease':
                cell.font = Font(name='Arial', size=9, bold=True, color='244185')
            else:
                cell.font = Font(name='Arial', size=9, color='595959')

# Auto column widths
for ci_col in range(1, len(cols)+1):
    max_len = max(len(str(ws.cell(r, ci_col).value or '')) for r in range(1, len(table_df)+2))
    ws.column_dimensions[get_column_letter(ci_col)].width = min(max_len + 2, 22)
ws.row_dimensions[1].height = 30
ws.freeze_panes = 'A2'

# ── Sheet 2: Model parameters ───────────────────────────────────────────────
ws2 = wb.create_sheet('Model Parameters')
param_rows = [
    ('Parameter', 'Value', 'Description'),
    ('Spatial block size', f'{GRID_KM} km', 'Grid cell size for spatial blocking'),
    ('CV folds (k)', K_FOLDS, 'Number of spatial cross-validation folds'),
    ('CV repetitions', N_REPS, 'Independent fold assignments averaged'),
    ('Bootstrap resamples', BOOT_B, 'Block bootstrap iterations for 95% CI'),
    ('Random Forest trees', N_TREES, 'Number of trees per RF model'),
    ('mtry', 'p/3', 'Features per split (p = no. predictors)'),
    ('Min. leaf size', MIN_LEAF, 'Minimum samples per leaf node'),
    ('Response transform', 'log1p(y)', 'Applied before model; expm1 for back-transform'),
    ('Counterfactual Δ', 'Pred(Upland) − Pred(Paddy)', 'Evaluated on paddy sites, OOF only'),
    ('CI method', 'Block bootstrap', 'Blocks = same 100-km grid as CV'),
    ('Random seed', SEED, 'Reproducibility seed'),
]
for ri, row in enumerate(param_rows, 1):
    for ci_col, val in enumerate(row, 1):
        cell = ws2.cell(row=ri, column=ci_col, value=val)
        cell.font = Font(name='Arial', size=9, bold=(ri == 1))
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border
        if ri == 1:
            cell.fill = header_fill
            cell.font = hdr_font
for ci_col, w in enumerate([28, 30, 48], 1):
    ws2.column_dimensions[get_column_letter(ci_col)].width = w

wb.save(xlsx_path)
print(f"    Saved: {xlsx_path}")

# ════════════════════════════════════════════════════════════════════════════════
# 6. NATURE-STYLE PUBLICATION FIGURE
# ════════════════════════════════════════════════════════════════════════════════
print("\n[4] Generating Nature-style figure...")

# Colour scheme
DEPTH_COL = {'0–30 cm': '#2166AC', '30–100 cm': '#B2182B'}
VAR_COL   = {'SOC': '#4D9221', 'SIC': '#C27D38'}
ERR_KW    = dict(fmt='none', capsize=3.5, capthick=0.8, lw=0.8, zorder=4)

fig = plt.figure(figsize=(183/25.4, 155/25.4))
gs  = gridspec.GridSpec(2, 2,
                         left=0.12, right=0.97,
                         top=0.92,  bottom=0.14,
                         wspace=0.35, hspace=0.52)

panel_letters = ['a', 'b', 'c', 'd']

# ── Panel a: Δ forest plot (horizontal) ─────────────────────────────────────
ax_forest = fig.add_subplot(gs[0, :])   # full width top row

y_labels, y_pos, colors_used = [], [], []
delta_vals, lo_vals, hi_vals = [], [], []
sig_vals = []

for yi, (resp, depth) in enumerate([
    ('SOC', '0–30 cm'), ('SOC', '30–100 cm'),
    ('SIC', '0–30 cm'), ('SIC', '30–100 cm'),
    ('TC', '0–30 cm'), ('TC', '30–100 cm'),
]):
    row = results_df[(results_df['Response'] == resp) & (results_df['Depth'] == depth)]
    if row.empty:
        continue
    row = row.iloc[0]
    lbl = f'{resp}\n{depth}'
    y_labels.append(lbl)
    y_pos.append(yi)
    colors_used.append(DEPTH_COL[depth])
    delta_vals.append(row['Mean_Delta'])
    lo_vals.append(row['CI_lo95'])
    hi_vals.append(row['CI_hi95'])
    sig_vals.append(row['Significance'])

y_pos = list(range(len(y_labels)))[::-1]   # top to bottom

for yi, (yp, dv, lo, hi, sc, lbl) in enumerate(
        zip(y_pos, delta_vals, lo_vals, hi_vals, sig_vals, y_labels)):
    c = colors_used[yi]
    # CI bar
    ax_forest.barh([yp], [0], color='none')   # placeholder for y range
    ax_forest.plot([lo, hi], [yp, yp], color=c, lw=2.2, solid_capstyle='round', zorder=3)
    # caps
    for xv in [lo, hi]:
        ax_forest.plot([xv, xv], [yp - 0.12, yp + 0.12], color=c, lw=1.4, zorder=4)
    # point estimate
    marker = ('D' if sc == 'Increase' else 'v' if sc == 'Decrease' else 'o')
    ax_forest.scatter([dv], [yp], color=c, s=45, zorder=5, marker=marker,
                      edgecolors='white', linewidths=0.6)
    # significance label
    slbl = ('***' if sc != 'ns' else 'ns')
    ax_forest.text(hi + 0.015, yp, slbl, va='center', ha='left',
                   fontsize=7.5, color=c, fontweight='bold')

ax_forest.axvline(0, color='#555', lw=0.8, linestyle='--', zorder=2)
ax_forest.set_yticks(y_pos)
ax_forest.set_yticklabels(y_labels, fontsize=8.5)
ax_forest.set_ylim(-0.5, len(y_labels) - 0.5)
ax_forest.margins(x=0.1)
ax_forest.set_xlabel(r'Counterfactual $\Delta$ (g kg$^{-1}$)', fontsize=9, labelpad=3)
ax_forest.set_title('Paddy → Upland counterfactual change (Δ = Pred(Upland) − Pred(Paddy))',
                     fontsize=9, pad=12, fontweight='bold')
ax_forest.spines['top'].set_visible(False)
ax_forest.spines['right'].set_visible(False)
ax_forest.xaxis.set_minor_locator(AutoMinorLocator(2))
ax_forest.tick_params(which='minor', length=2, width=0.4, direction='in')
ax_forest.grid(axis='x', linewidth=0.3, linestyle=':', color='#ccc', zorder=0)
ax_forest.set_axisbelow(True)

# Custom legend
legend_elems = [
    mpatches.Patch(color=DEPTH_COL['0–30 cm'],   label='0–30 cm'),
    mpatches.Patch(color=DEPTH_COL['30–100 cm'], label='30–100 cm'),
    plt.scatter([], [], marker='D', c='#555', s=30, label='Significant increase'),
    plt.scatter([], [], marker='o', c='#555', s=30, label='Not significant'),
]
ax_forest.legend(handles=legend_elems, loc='center left', frameon=True,
                 framealpha=0.9, edgecolor='#ccc', fontsize=7.5,
                 handlelength=1.2, labelspacing=0.4)
ax_forest.text(-0.09, 1.06, 'a', transform=ax_forest.transAxes,
               fontsize=11, fontweight='bold', va='top', ha='left')

# ── Panel b: CV R² (original scale) ─────────────────────────────────────────
ax_r2 = fig.add_subplot(gs[1, 0])

bar_labels = [f"{r['Response']}\n{r['Depth']}" for _, r in results_df.iterrows()]
bar_colors = [DEPTH_COL[r['Depth']] for _, r in results_df.iterrows()]
r2_vals    = results_df['CV_R2_orig'].values
x_pos      = np.arange(len(bar_labels))

bars = ax_r2.bar(x_pos, r2_vals, color=bar_colors, alpha=0.85,
                  width=0.6, edgecolor='white', linewidth=0.5)
for xi, (b, rv) in enumerate(zip(bars, r2_vals)):
    ax_r2.text(xi, rv + 0.01, f'{rv:.2f}', ha='center', va='bottom',
               fontsize=7.5, color='#333')

ax_r2.set_xticks(x_pos)
ax_r2.set_xticklabels(bar_labels, fontsize=8)
ax_r2.set_ylabel(r'CV $R^2$ (original scale)', fontsize=9, labelpad=3)
ax_r2.set_title('Model performance (CV $R^2$)', fontsize=9, pad=5, fontweight='bold')
ax_r2.set_ylim(0, max(r2_vals) * 1.25)
ax_r2.spines['top'].set_visible(False)
ax_r2.spines['right'].set_visible(False)
ax_r2.yaxis.set_minor_locator(AutoMinorLocator(2))
ax_r2.tick_params(which='minor', length=2, width=0.4, direction='in')
ax_r2.yaxis.grid(linewidth=0.3, linestyle=':', color='#ccc', zorder=0)
ax_r2.set_axisbelow(True)
ax_r2.text(-0.18, 1.06, 'b', transform=ax_r2.transAxes,
           fontsize=11, fontweight='bold', va='top', ha='left')

# ── Panel c: CV RMSE (original scale) ───────────────────────────────────────
ax_rmse = fig.add_subplot(gs[1, 1])

rmse_vals = results_df['CV_RMSE_orig'].values
bars2 = ax_rmse.bar(x_pos, rmse_vals, color=bar_colors, alpha=0.85,
                     width=0.6, edgecolor='white', linewidth=0.5)
for xi, (b, rv) in enumerate(zip(bars2, rmse_vals)):
    ax_rmse.text(xi, rv + max(rmse_vals)*0.01, f'{rv:.2f}', ha='center', va='bottom',
                 fontsize=7.5, color='#333')

ax_rmse.set_xticks(x_pos)
ax_rmse.set_xticklabels(bar_labels, fontsize=8)
ax_rmse.set_ylabel(r'CV RMSE (g kg$^{-1}$)', fontsize=9, labelpad=3)
ax_rmse.set_title(r'Model performance (CV RMSE)', fontsize=9, pad=5, fontweight='bold')
ax_rmse.set_ylim(0, max(rmse_vals) * 1.25)
ax_rmse.spines['top'].set_visible(False)
ax_rmse.spines['right'].set_visible(False)
ax_rmse.yaxis.set_minor_locator(AutoMinorLocator(2))
ax_rmse.tick_params(which='minor', length=2, width=0.4, direction='in')
ax_rmse.yaxis.grid(linewidth=0.3, linestyle=':', color='#ccc', zorder=0)
ax_rmse.set_axisbelow(True)
ax_rmse.text(-0.18, 1.06, 'c', transform=ax_rmse.transAxes,
             fontsize=11, fontweight='bold', va='top', ha='left')

# ── Subtitle/footnote ────────────────────────────────────────────────────────
fig.text(0.50, 0.02,
         f'Spatial block CV (k={K_FOLDS}, {N_REPS} reps, {GRID_KM}-km blocks, {N_TREES} trees). '
         r'95% CI via block bootstrap ($B$=' + str(BOOT_B) + r').',
         ha='center', va='bottom', fontsize=6.8, color='#555', style='italic')

# ── Save ─────────────────────────────────────────────────────────────────────
png_path = f'{OUT_DIR}\\SpatialCV_RF_CounterfactualDelta.png'
pdf_path = f'{OUT_DIR}\\SpatialCV_RF_CounterfactualDelta.pdf'
fig.savefig(png_path, dpi=600, bbox_inches='tight', format='png')
fig.savefig(pdf_path, bbox_inches='tight', format='pdf')
print(f"    Saved: {png_path}")
print(f"    Saved: {pdf_path}")

print("\n" + "="*70)
print("Done. All outputs saved to:", OUT_DIR)
print("="*70)
